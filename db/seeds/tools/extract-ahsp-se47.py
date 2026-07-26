# Ekstraktor AHSP SE 47/2026 → dataset kanonik (db/seeds/ahsp-se47-dataset.json)
#
# Zero-Invention (ADR-006): salin VERBATIM dari workbook resmi; tanpa tafsir angka.
# Jalankan lokal (workbook di _source/, TIDAK di git):
#   python db/seeds/tools/extract-ahsp-se47.py
#
# Aturan:
# - Blok analisa: kolom C kode (\d+(\.\d+)+), kolom D uraian; komponen di bawah
#   section A TENAGA KERJA / B BAHAN / C PERALATAN; berhenti di baris F.
# - Satuan: case-fold + peta eksplisit; TANPA tafsir semantik. Typo workbook
#   (loat/lkg) dipertahankan sebagai kode satuan tersendiri (baseline "SE bilang apa").
# - Satuan output: parse '1 <sat>' / 'per <sat>' dari uraian (verbatim di teks;
#   'tititk' = typo titik di teks SE, ikut ditangkap). Tak terparse → EXCLUDED,
#   masuk dataset.excluded utk ditinjau — TIDAK ditebak.
# - Resource: identitas EXACT (nama case-insensitive + satuan). Tanpa fuzzy-merge.
import json, re, hashlib, sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]  # tools → seeds → db → ROOT repo
SRC = ROOT / '_source' / 'ahsp' / 'AHSP CIPTA KARYA SE BINA KONTRUKSI NO. 47 TAHUN 2026.xlsm'
OUT = ROOT / 'db' / 'seeds' / 'ahsp-se47-dataset.json'

SKIP_SHEETS = {'Upah Bahan', 'Daftar Harga Satuan Pekerjaan', 'Pembesian Plat Lantai 1'}
CODE_RE = re.compile(r'^\d+(\.\d+)+$')

EXISTING_UNITS = {'oh': 'OH', 'oj': 'OJ', 'jam': 'jam', 'hari': 'hari', 'm': 'm', 'm2': 'm2',
  'm3': 'm3', 'kg': 'kg', 'buah': 'buah', 'bh': 'buah', 'unit': 'unit', 'liter': 'liter',
  'lembar': 'lembar', 'batang': 'batang', 'set': 'set', 'ton': 'ton', 'ls': 'ls',
  'sak': 'sak', 'titik': 'titik', 'rol': 'rol', 'minggu': 'minggu'}
NEW_UNITS = {"m'": 'm1', 'lot': 'lot', 'tube': 'tube', 'gulung': 'gulung', 'pohon': 'pohon',
  'polybag': 'polybag', 'ikat': 'ikat', 'dus': 'dus', 'cm': 'cm',
  'unit hari': 'unit_hari', 'buah hari': 'buah_hari', 'unit/hari': 'unit_hari',
  'loat': 'loat', 'lkg': 'lkg'}

def map_unit(raw: str):
    k = raw.strip().lower()
    return EXISTING_UNITS.get(k) or NEW_UNITS.get(k)

OUT1 = re.compile(r"\b1\s*(m3|m2|m'|m1|m|kg|buah|bh|unit|titik|tititk|ls|set|lembar|batang|zak|ton|liter)\b", re.I)
OUT2 = re.compile(r"\bper\s+(?:1\s+)?(m3|m2|m'|m1|m|kg|buah|unit|titik)\b", re.I)

def output_unit(uraian: str):
    u = uraian.lower().replace('m³', 'm3').replace('m²', 'm2')
    m = OUT1.search(u) or OUT2.search(u)
    if not m:
        return None
    tok = m.group(1).replace('tititk', 'titik')  # typo verbatim SE (2.4.5.6)
    return map_unit(tok) or tok

def extract():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    raw = []
    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS:
            continue
        ws = wb[sname]
        cur, section = None, None
        for row in ws.iter_rows(min_row=1, max_col=9, values_only=True):
            c = str(row[2]).strip() if row[2] is not None else ''
            d = str(row[3]).strip() if row[3] is not None else ''
            if CODE_RE.match(c) and d:
                cur = {'sheet': sname, 'code': c, 'uraian': d, 'components': []}
                raw.append(cur); section = None
                continue
            if cur is None:
                continue
            if c == 'A' or d.upper().startswith('TENAGA KERJA'):
                section = 'labor'; continue
            if c == 'B' or d.upper() == 'BAHAN':
                section = 'material'; continue
            if c == 'C' or d.upper() == 'PERALATAN':
                section = 'equipment'; continue
            if isinstance(row[2], str) and row[2].startswith('JUMLAH'):
                continue
            if c in ('D', 'E', 'F'):
                # cacat blok-sum: D workbook != Σ jumlah baris komponen (range SUM
                # workbook salah/lompat baris) → F-nya mewarisi selisih. Koefisien
                # kolom G tetap normatif; ini didokumentasikan, bukan diikuti.
                if c == 'D' and isinstance(row[8], (int, float)):
                    sumj = sum(x.get('_jumlah', 0.0) for x in cur['components'])
                    if abs(sumj - float(row[8])) > max(1.0, 1e-6 * abs(float(row[8]))):
                        cur.setdefault('defects', []).append(
                            f'blok-sum: D workbook={float(row[8]):.6g} != Σ jumlah komponen={sumj:.6g} '
                            f'(range SUM workbook cacat; F workbook mewarisi selisih)')
                if c == 'F':
                    cur = None
                continue
            if d and section and row[6] is not None:
                koef, note = None, None
                if isinstance(row[6], (int, float)):
                    koef = float(row[6])
                else:
                    # cacat workbook: koef TEKS (mis. '2,625'). Disambiguasi lewat
                    # perilaku workbook SENDIRI: jumlah/harga (kalau keduanya angka).
                    if isinstance(row[7], (int, float)) and isinstance(row[8], (int, float)) and row[7]:
                        koef = float(row[8]) / float(row[7])
                        note = f'koef teks {str(row[6])!r}; dipakai jumlah/harga workbook = {koef}'
                    else:
                        cur.setdefault('defects', []).append(
                            f'komponen "{d[:40]}" koef teks {str(row[6])!r} + jumlah rusak — komponen dilewati')
                        continue
                if koef is None:
                    continue
                sat = str(row[5]).strip() if row[5] not in (None, '') else None  # None → resolve pass-2
                comp = {'name': d, 'sat': sat, 'koef': koef, 'section': section}
                if isinstance(row[8], (int, float)):
                    comp['_jumlah'] = float(row[8])   # utk cek blok-sum D (tak ikut dataset)
                if note:
                    comp['note'] = note
                # cacat workbook: jumlah != koef*harga (formula salah-baris dsb) — catat saja,
                # koefisien kolom G tetap normatif (verbatim).
                if (isinstance(row[6], (int, float)) and isinstance(row[7], (int, float))
                        and isinstance(row[8], (int, float)) and row[7]):
                    eff = float(row[8]) / float(row[7])
                    if abs(eff - koef) / max(abs(koef), 1e-12) > 1e-6:
                        cur.setdefault('defects', []).append(
                            f'komponen "{d[:40]}": jumlah workbook memakai {eff:.6g}, koef kolom = {koef:.6g}')
                cur['components'].append(comp)
    wb.close()
    return [a for a in raw if a['components']]

def main():
    raw = extract()
    # ── pass-2: komponen ber-sat KOSONG (cacat workbook) → resolve dari nama yang
    #    sama di tempat lain workbook bila satuannya UNIK (derivasi mekanis dari
    #    sumber yang sama, bukan tafsiran luar). Ambigu/tak-ada → komponen defect.
    sat_by_name = {}
    for s in raw:
        for cmp in s['components']:
            if cmp['sat']:
                sat_by_name.setdefault(cmp['name'].strip().lower(), set()).add(cmp['sat'])
    for s in raw:
        for cmp in list(s['components']):
            if cmp['sat'] is None:
                raw_c = sat_by_name.get(cmp['name'].strip().lower(), set())
                cands = {c for c in ((map_unit(x) or x.strip().lower()) for x in raw_c)}
                if len(cands) == 1:
                    cmp['sat'] = next(iter(cands))
                    cmp['note'] = (cmp.get('note', '') + '; ' if cmp.get('note') else '') + \
                        f'sat kosong di workbook — dipakai satuan nama-sama di blok lain: {cmp["sat"]}'
                else:
                    s.setdefault('defects', []).append(
                        f'komponen "{cmp["name"][:40]}" tanpa satuan & tak ter-resolve (kandidat: {sorted(cands)}) — dilewati')
                    s['components'].remove(cmp)

    res_index, resources = {}, []
    analyses, excluded = [], []
    seen_codes = {}   # code → uraian pertama (deteksi kode ganda DI DALAM workbook)
    for s in raw:
        # ── cacat workbook: kode analisa GANDA ──
        if s['code'] in seen_codes:
            if s['uraian'].strip().lower() == seen_codes[s['code']].strip().lower():
                # duplikat identik (mis. 6.5.5.5) → satu saja, catat
                excluded.append({'code': s['code'], 'sheet': s['sheet'],
                                 'reason': 'duplikat identik di workbook (baris kedua dilewati)'})
            else:
                # kode sama, item BERBEDA (mis. 6.5.5.4 Floater vs Foot Valve) —
                # identitas bentrok; TIDAK ditebak pelabelannya → excluded utk keputusan founder
                excluded.append({'code': s['code'], 'sheet': s['sheet'],
                                 'uraian': s['uraian'],
                                 'reason': 'kode GANDA di workbook untuk item BERBEDA — butuh keputusan pelabelan'})
            continue
        seen_codes[s['code']] = s['uraian']

        ou = output_unit(s['uraian'])
        comps_by_res, order, badunit, dupnotes = {}, [], [], []
        for cmp in s['components']:
            sat = map_unit(cmp['sat'])
            if not sat:
                badunit.append(cmp['sat']); continue
            key = (cmp['name'].strip().lower(), sat)
            if key not in res_index:
                code = f'AHSP-R{len(resources)+1:04d}'
                res_index[key] = code
                resources.append({'code': code, 'name': cmp['name'].strip(),
                                  'unit_code': sat, 'category': cmp['section']})
            rcode = res_index[key]
            if rcode in comps_by_res:
                # cacat workbook: resource sama 2 baris dlm satu analisa →
                # JUMLAHKAN koefisien (Σk×p == (Σk)×p — paritas eksak) + provenance
                prev = comps_by_res[rcode]['k']
                comps_by_res[rcode]['k'] = prev + cmp['koef']
                dupnotes.append(f"{cmp['name'].strip()}: workbook 2 baris ({prev} + {cmp['koef']})")
            else:
                comps_by_res[rcode] = {'r': rcode, 'k': cmp['koef']}
                order.append(rcode)
        if badunit:
            excluded.append({'code': s['code'], 'sheet': s['sheet'],
                             'reason': f'satuan tak terpeta: {sorted(set(badunit))}'})
            continue
        if not ou:
            excluded.append({'code': s['code'], 'sheet': s['sheet'],
                             'reason': 'satuan output tak tertulis di uraian'})
            continue
        rec = {'code': s['code'], 'sheet': s['sheet'], 'uraian': s['uraian'],
               'output_unit': ou, 'components': [comps_by_res[r] for r in order]}
        allnotes = dupnotes + [f"{c['name'][:40]}: {c['note']}" for c in s['components'] if c.get('note')]
        if allnotes:
            rec['notes'] = '; '.join(allnotes)
        if s.get('defects'):
            rec['workbook_defects'] = s['defects']
        analyses.append(rec)

    sha = hashlib.sha256(SRC.read_bytes()).hexdigest()
    dataset = {
        'meta': {
            'source_file': SRC.name, 'source_sha256': sha, 'edition_code': 'SE-47-2026',
            'extractor': 'db/seeds/tools/extract-ahsp-se47.py',
            'counts': {'analyses': len(analyses), 'resources': len(resources),
                       'components': sum(len(a['components']) for a in analyses),
                       'excluded': len(excluded)},
        },
        'new_units': sorted(set(NEW_UNITS.values())),
        'resources': resources,
        'analyses': analyses,
        'excluded': excluded,
    }
    OUT.write_text(json.dumps(dataset, ensure_ascii=False), encoding='utf-8')
    print(f"OK → {OUT}")
    print(json.dumps(dataset['meta']['counts']))

if __name__ == '__main__':
    sys.exit(main())
