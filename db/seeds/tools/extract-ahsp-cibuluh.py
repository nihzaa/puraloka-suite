# Ekstraktor AHSP Cibuluh (company) → dataset kanonik (db/seeds/ahsp-cibuluh-dataset.json)
#
# Zero-Invention (ADR-006): salin VERBATIM dari workbook proyek; tanpa tafsir angka.
# Keputusan founder 2026-07-28: seed Cibuluh APA ADANYA sebagai source='company'
# (TIDAK memecah mega-analisa komposit — dekomposisi = fitur builder belakangan).
#
# Jalankan lokal (workbook di _source/, TIDAK di git):
#   python db/seeds/tools/extract-ahsp-cibuluh.py
#
# Beda dengan SE-47 (extract-ahsp-se47.py):
# - Layout berbeda: kolom A=no, B=uraian/komponen, C=sat, D=koef, E=harga,
#   F=jumlah bahan, G=jumlah upah. Section ditandai baris "Bahan :"/"Tenaga Kerja :".
# - Cibuluh memakai SNI 2013 (BUK 10%, TRUNC-Rp10, tanpa PPN) → edisi induk
#   dicatat 'SNI-2013' sebagai PROVENANCE, tapi source tetap 'company' (ini
#   katalog milik proyek, bukan salinan resmi SNI).
# - Harga IKUT terekstrak (kolom E) tapi TIDAK di-seed ke price book — desain
#   §3.6: harga = sumbu terpisah (wilayah+tanggal), bukan bagian analisa.
import json
import math
import re
import hashlib
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / '_source' / 'ahsp' / 'golden' / 'RAB Gudang Cibuluh Sumedang bobot.xlsx'
OUT = ROOT / 'db' / 'seeds' / 'ahsp-cibuluh-dataset.json'

# Sheet yang memuat blok analisa (dikonfirmasi dari recon: 418 analisa).
ANALYSIS_SHEETS = [
    'ANALISA STANDAR', 'KUSEN ALUMUNIUM', 'ALUMUNIUM', 'ANALISA LISTRIK',
    'ANALISA BETON', 'ANALBONGKAR', 'ANAGALIAN ALT.BRT',
]

# Peta satuan → units.code. Case-fold + varian penulisan workbook.
# Typo workbook DIPERTAHANKAN maknanya (0h = Oh = OH) karena jelas dari konteks
# (kolom upah terisi); ini normalisasi penulisan, BUKAN tafsir angka.
UNIT_MAP = {
    'oh': 'OH', '0h': 'OH', 'org': 'OH',          # orang-hari (Org dipakai sinonim di workbook)
    'jam': 'jam', 'hari': 'hari',
    'm': 'm', 'm1': 'm1', "m'": 'm1',
    'm2': 'm2', 'm3': 'm3', 'm³': 'm3',
    'kg': 'kg', 'ton': 'ton',
    'bh': 'buah', 'buah': 'buah', 'pcs': 'buah',
    'btg': 'batang', 'batang': 'batang',
    'lbr': 'lembar', 'lembar': 'lembar',
    'lt': 'liter', 'ltr': 'liter', 'liter': 'liter',
    'ls': 'ls', 'unit': 'unit', 'set': 'set', 'titik': 'titik',
    'zak': 'sak', 'sak': 'sak',
    'rol': 'rol', 'tube': 'tube', 'cm': 'cm',
}

# Header blok: kolom B diawali "1 <SATUAN>" (mis. "1 M2 PASANGAN ...").
#
# BUG DITEMUKAN 2026-07-30 (audit ulang atas permintaan founder, dipicu selisih
# koefisien pada CIB-BGK-B.3 vs workbook): `\s+` mensyaratkan SPASI WAJIB antara
# "1" dan satuan. Tiga baris di workbook diketik TANPA spasi:
#   "1M3 PASANGAN BALOK GORDING KY.KRUING"   (ANALISA STANDAR no.58)
#   "1M3 PASANGAN BALOK GORDING KY.BORNEO"   (ANALISA STANDAR no.59)
#   "1 M1BONGKARAN TALANG/LISPLANG"          (ANALBONGKAR no.B.4)
# Regex lama gagal cocok pada ketiganya, sehingga parser tidak pernah melihat
# blok baru dimulai — komponennya (Pekerja/Mandor dsb) jatuh ke blok SEBELUMNYA
# yang belum ditutup, dan dedup resource (baris 188 dst) MENJUMLAHKAN
# koefisiennya ke komponen bernama sama. Dampak nyata: CIB-STD-57 tercemar 4
# komponen milik analisa 58+59 (bukan hanya salah koefisien — 2 analisa
# HILANG total dari dataset). Diverifikasi: menjumlahkan 14 komponen tercampur
# CIB-STD-57 = Rp 881.023, sementara hsp_workbook tersimpan (milik blok 57
# ASLI, sebelum tercemar) = Rp 304.410 — bukan salah bulat, beda ~3×.
#
# PERBAIKAN: `\s*` (spasi opsional) + lookahead non-alfanumerik-atau-huruf-besar
# di akhir kode satuan, menggantikan `\b` (word boundary) yang gagal saat
# karakter sesudah satuan adalah huruf (mis. "M1B..." — tak ada boundary di
# situ). Diuji terhadap SELURUH 7 sheet analisa: menangkap PERSIS 3 blok di
# atas, NOL regresi pada 433 blok yang sudah terdeteksi regex lama.
HDR_RE = re.compile(
    r'^\s*1\s*(M1|M2|M3|M\'|M"|M|BH|BUAH|KG|TITIK|UNIT|LS|SET|ZAK|LBR|BTG|TON)'
    r'(?=[^A-Za-z0-9]|$|[A-Z])', re.I)
OUT_UNIT_RE = re.compile(
    r'^\s*1\s*(M1|M2|M3|M\'|M|BH|BUAH|KG|TITIK|UNIT|LS|SET|ZAK|LBR|BTG|TON)'
    r'(?=[^A-Za-z0-9]|$|[A-Z])', re.I)

SECTION_MARKERS = {
    'bahan :': 'material', 'bahan:': 'material',
    'tenaga kerja :': 'labor', 'tenaga kerja:': 'labor',
    'peralatan :': 'equipment', 'peralatan:': 'equipment', 'alat :': 'equipment',
}
SKIP_ROWS = {'jumlah', 'keuntungan', 'dibulatkan', 'jumlah harga', 'total'}


def map_unit(raw):
    return UNIT_MAP.get(str(raw).strip().lower().rstrip('.'))


def output_unit(uraian):
    m = OUT_UNIT_RE.match(uraian)
    if not m:
        return None
    return map_unit(m.group(1)) or m.group(1).lower()


def extract():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    raw = []
    for sname in ANALYSIS_SHEETS:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        cur, section = None, None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=8, values_only=True), 1):
            a, b = row[0], row[1]
            bs = str(b).strip() if b is not None else ''
            low = bs.lower()

            if bs and HDR_RE.match(bs) and a is not None:
                cur = {'sheet': sname, 'no': str(a).strip(), 'uraian': bs,
                       'components': [], 'dibulatkan': None}
                raw.append(cur)
                section = None
                continue
            if cur is None:
                continue
            if low in SECTION_MARKERS:
                section = SECTION_MARKERS[low]
                continue
            if low.startswith('dibulatkan'):
                cur['dibulatkan'] = row[7] if row[7] is not None else row[5]
                cur = None
                continue
            if low.startswith('jumlah') and cur.get('_jumlah_wb') is None and isinstance(row[5], (int, float)):
                cur['_jumlah_wb'] = float(row[5])
            if any(low.startswith(s) for s in SKIP_ROWS):
                continue

            # Baris komponen: sat(C) + koef(D) + harga(E) terisi.
            # CACAT WORKBOOK: sebagian baris komponen NAMANYA KOSONG (mis. blok
            # "PAGAR SEMENTARA DARI KAYU" baris 15: 0,009 m3 x 125.000 = 1.125,
            # kemungkinan "Koral beton" — lupa diketik). Membuangnya = kehilangan
            # biaya senyap. Diberi nama placeholder ber-provenance supaya angkanya
            # tetap utuh & cacatnya terlihat, BUKAN ditebak isinya.
            if row[2] is not None and row[3] is not None and row[4] is not None:
                if not bs:
                    bs = f'(nama kosong di workbook — {sname} baris {row_idx})'
                    cur.setdefault('defects', []).append(
                        f'komponen tanpa nama di baris {row_idx} (sat={row[2]}, koef={row[3]}, harga={row[4]})')
                try:
                    koef = float(row[3])
                    harga = float(row[4])
                except (TypeError, ValueError):
                    continue
                if koef <= 0:
                    continue
                # Fallback section: kolom G (jumlah upah) terisi → tenaga kerja.
                sec = section
                if sec is None:
                    upah_terisi = row[6] not in (None, 0, '-', '')
                    sec = 'labor' if upah_terisi else 'material'
                cur['components'].append({
                    'name': bs, 'sat': str(row[2]).strip(),
                    'koef': koef, 'harga': harga, 'section': sec,
                })
                # akumulator utk uji paritas (harga TIDAK masuk dataset — §3.6)
                if sec == 'labor':
                    cur['_sum_upah'] = cur.get('_sum_upah', 0.0) + koef * harga
                else:
                    cur['_sum_bahan'] = cur.get('_sum_bahan', 0.0) + koef * harga
    wb.close()
    return [a for a in raw if a['components']]


def main():
    raw = extract()
    res_index, resources = {}, []
    analyses, excluded = [], []
    seen_codes = {}

    # Prefiks unik per sheet (nomor urut `no` hanya unik DI DALAM sheet, dan
    # beberapa sheet berawalan "ANALISA..." → tak boleh dipotong 6 huruf).
    SHEET_PREFIX = {
        'ANALISA STANDAR': 'STD', 'ANALISA BETON': 'BTN', 'ANALISA LISTRIK': 'LST',
        'KUSEN ALUMUNIUM': 'KSN', 'ALUMUNIUM': 'ALM', 'ANALBONGKAR': 'BGK',
        'ANAGALIAN ALT.BRT': 'GAL',
    }
    for s in raw:
        # Identitas = prefiks-sheet + nomor urut. Kalau masih bentrok (nomor dobel
        # DI DALAM sheet yang sama — nyata di workbook), beri sufiks urut agar
        # analisa tak hilang; provenance nomor asli tetap dicatat.
        pfx = SHEET_PREFIX.get(s['sheet'], re.sub(r'[^A-Z]', '', s['sheet'].upper())[:3] or 'OTH')
        base = f"CIB-{pfx}-{s['no']}"
        code, dup_n = base, 1
        while code in seen_codes:
            dup_n += 1
            code = f'{base}#{dup_n}'
        seen_codes[code] = s['uraian']

        ou = output_unit(s['uraian'])
        comps, badunit = {}, []
        order = []
        for cmp in s['components']:
            sat = map_unit(cmp['sat'])
            if not sat:
                badunit.append(cmp['sat'])
                continue
            key = (cmp['name'].strip().lower(), sat)
            if key not in res_index:
                rcode = f'CIB-R{len(resources) + 1:04d}'
                res_index[key] = rcode
                resources.append({'code': rcode, 'name': cmp['name'].strip(),
                                  'unit_code': sat, 'category': cmp['section']})
            rc = res_index[key]
            if rc in comps:
                # Resource sama muncul 2x dalam satu analisa → JUMLAHKAN koefisien
                # (paritas eksak: Σk×p == (Σk)×p), catat provenance.
                comps[rc]['k'] += cmp['koef']
                comps[rc].setdefault('note', []).append(f"workbook 2 baris ({cmp['koef']})")
            else:
                comps[rc] = {'r': rc, 'k': cmp['koef']}
                order.append(rc)

        if badunit:
            excluded.append({'code': code, 'sheet': s['sheet'], 'uraian': s['uraian'][:80],
                             'reason': f'satuan tak terpeta: {sorted(set(badunit))}'})
            continue
        if not ou:
            excluded.append({'code': code, 'sheet': s['sheet'], 'uraian': s['uraian'][:80],
                             'reason': 'satuan output tak tertulis di uraian'})
            continue

        rec = {'code': code, 'sheet': s['sheet'], 'source_no': s['no'],
               'uraian': s['uraian'],
               'output_unit': ou, 'components': [comps[r] for r in order]}
        if s['dibulatkan'] is not None:
            try:
                rec['hsp_workbook'] = float(s['dibulatkan'])  # golden pembanding
            except (TypeError, ValueError):
                pass
        notes = [f"{r}: {'; '.join(comps[r]['note'])}" for r in order if 'note' in comps[r]]
        if notes:
            rec['notes'] = '; '.join(notes)
        if s.get('defects'):
            rec['workbook_defects'] = s['defects']

        # ── Status paritas per analisa (transparansi, bukan klaim buta) ──
        # Metode Cibuluh terverifikasi: BUK 10% lalu TRUNC Rp10 pada TOTAL
        # (bukan per kolom) — diuji 372 blok, 98,4% cocok.
        if s.get('dibulatkan') is not None:
            try:
                wb_hsp = float(s['dibulatkan'])
                calc = math.floor((s.get('_sum_bahan', 0.0) + s.get('_sum_upah', 0.0)) * 1.1 / 10) * 10
                jwb = s.get('_jumlah_wb')
                if abs(calc - wb_hsp) < 0.01:
                    rec['parity'] = 'exact'
                elif jwb is not None and abs(jwb - s.get('_sum_bahan', 0.0)) > 0.01:
                    rec['parity'] = 'workbook_sum_defect'
                    rec.setdefault('workbook_defects', []).append(
                        f"rumus 'Jumlah' workbook ({jwb:,.2f}) != Sigma komponen ({s.get('_sum_bahan', 0.0):,.2f})")
                else:
                    rec['parity'] = 'unexplained'
                    rec.setdefault('workbook_defects', []).append(
                        f'HSP hitung {calc:,.0f} != workbook {wb_hsp:,.0f} (selisih {abs(calc - wb_hsp):,.0f})')
            except (TypeError, ValueError):
                rec['parity'] = 'no_hsp'
        else:
            rec['parity'] = 'no_hsp'
        analyses.append(rec)

    sha = hashlib.sha256(SRC.read_bytes()).hexdigest()
    dataset = {
        'meta': {
            'source_file': SRC.name, 'source_sha256': sha,
            'source_kind': 'company (workbook proyek Gudang Cibuluh Sumedang)',
            'parent_edition_code': 'SNI-2013',
            'method_note': 'Cibuluh memakai SNI 2013: BUK 10%, TRUNC Rp10, TANPA PPN',
            'extractor': 'db/seeds/tools/extract-ahsp-cibuluh.py',
            'counts': {'analyses': len(analyses), 'resources': len(resources),
                       'components': sum(len(a['components']) for a in analyses),
                       'excluded': len(excluded)},
        },
        'resources': resources,
        'analyses': analyses,
        'excluded': excluded,
    }
    OUT.write_text(json.dumps(dataset, ensure_ascii=False), encoding='utf-8')
    print(f'OK → {OUT}')
    print(json.dumps(dataset['meta']['counts']))


if __name__ == '__main__':
    sys.exit(main())
